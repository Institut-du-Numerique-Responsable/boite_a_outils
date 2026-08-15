#!/usr/bin/env python3
"""Convertit le classeur de travail INR en outils_raw.json.

Usage :
    python3 tools/import_xlsx.py "chemin/vers/Boite_a_outils_NR.xlsx"

L'onglet lu est OUTILS-NR. Les en-têtes attendus sont ceux de la version 2.1
(avril 2024) ; une colonne renommée dans le classeur doit l'être ici aussi.
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl est requis : python3 -m pip install openpyxl")

ONGLET = "OUTILS-NR"
HERE = os.path.dirname(os.path.abspath(__file__))
SORTIE = os.path.join(HERE, "outils_raw.json")


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    chemin = sys.argv[1]

    classeur = openpyxl.load_workbook(chemin, data_only=True)
    if ONGLET not in classeur.sheetnames:
        sys.exit(f"Onglet « {ONGLET} » introuvable. Onglets présents : {classeur.sheetnames}")

    feuille = classeur[ONGLET]
    lignes = list(feuille.iter_rows(values_only=True))
    entetes = [(h or "").strip() if isinstance(h, str) else h for h in lignes[0]]

    outils = []
    for ligne in lignes[1:]:
        enregistrement = {
            entetes[i]: ligne[i]
            for i in range(len(entetes))
            if entetes[i]
        }
        if any(valeur not in (None, "") for valeur in enregistrement.values()):
            outils.append(enregistrement)

    with open(SORTIE, "w", encoding="utf-8") as fichier:
        json.dump(outils, fichier, ensure_ascii=False, indent=1, default=str)

    avec_lien = sum(1 for o in outils if o.get("Lien"))
    print(f"{len(outils)} lignes lues ({avec_lien} avec lien) → {SORTIE}")
    print("Étapes suivantes : python3 tools/audit_links.py puis python3 tools/build.py")


if __name__ == "__main__":
    main()
